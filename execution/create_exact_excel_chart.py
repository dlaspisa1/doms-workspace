"""
Create PowerPoint chart with EXACT positioning from Excel file
"""

import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from openpyxl import load_workbook
import os

# Beautiful red and blue color scheme
COLORS = {
    'header_bg': '#1E3A8A',      # Deep blue
    'header_text': 'white',
    'range_1m_bg': '#DC2626',    # Bold red for $1M
    'range_500k_bg': '#2563EB',  # Bright blue for $500k
    'range_250k_bg': '#7C3AED',  # Purple for $250k
    'range_text': 'white',
    'cell_border': '#E5E7EB',
    'cell_text': '#1F2937',
    'legend_bg': '#F3F4F6',
    'arrow_color': '#6B7280'
}

def create_exact_chart():
    """
    Read Excel cell-by-cell and recreate exact positioning
    """
    # Load Excel file
    wb = load_workbook('.tmp/email_attachments/Savings overview01.xlsx')
    ws = wb.active

    # Read exact cell values and positions
    print("Reading Excel file cell-by-cell...")

    # Create mapping of all cells
    cells = {}
    for row in range(1, 17):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                col_letter = chr(64 + col)  # A=65, B=66, etc
                cells[f"{col_letter}{row}"] = str(cell.value).strip()
                print(f"  {col_letter}{row}: {cell.value}")

    # Create figure
    fig, ax = plt.subplots(figsize=(20, 14))
    ax.set_xlim(0, 8)
    ax.set_ylim(0, 17)
    ax.axis('off')

    # Title
    ax.text(4, 16.5, 'Recommendations',
           ha='center', va='top',
           fontsize=22, fontweight='bold', color=COLORS['header_bg'])

    # Define column positions (matching Excel columns)
    col_positions = {
        'B': 0.5,   # Range labels
        'C': 1.4,   # Off-load Risk
        'D': 2.8,   # Improve purchasing
        'E': 4.2,   # Shift costs
        'F': 5.6,   # Reduce exposure
        'G': 7.0    # Clinical/utilization
    }

    col_width = 1.3
    row_height = 0.8

    # Draw column headers (row 1, columns C-G)
    header_y = 15.5
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        if f"{col_letter}1" in cells:
            x = col_positions[col_letter]
            # Header background
            rect = Rectangle((x, header_y - 0.6), col_width, 0.6,
                           facecolor=COLORS['header_bg'],
                           edgecolor='white', linewidth=2)
            ax.add_patch(rect)
            # Header text
            header_text = cells[f"{col_letter}1"]
            ax.text(x + col_width/2, header_y - 0.3,
                   header_text, ha='center', va='center',
                   fontsize=9, fontweight='bold', color=COLORS['header_text'],
                   wrap=True)

    # Draw grid and content (rows 2-14)
    # Create row backgrounds for each range
    range_rows = {
        '$1M': (2, 6),    # Rows 2-6
        '$500k': (7, 10), # Rows 7-10
        '$250k': (11, 13) # Rows 11-13
    }

    # Draw cells for rows 2-14
    for row in range(2, 15):
        y_pos = 15.5 - (row * row_height)

        # Draw cells for columns C-G
        for col_letter in ['C', 'D', 'E', 'F', 'G']:
            x = col_positions[col_letter]

            # Determine cell background color based on content and row
            cell_ref = f"{col_letter}{row}"
            has_content = cell_ref in cells

            # Subtle alternating colors with blue tint for filled cells
            if has_content:
                cell_bg = '#EFF6FF'  # Light blue for cells with content
            else:
                cell_bg = '#FAFAFA'  # Very light gray for empty cells

            # Draw cell border
            rect = Rectangle((x, y_pos - row_height), col_width, row_height,
                           facecolor=cell_bg,
                           edgecolor=COLORS['cell_border'], linewidth=1.5)
            ax.add_patch(rect)

            # Add cell content if exists
            if has_content:
                cell_text = cells[cell_ref]
                # Add bullet point if not a header
                if row > 1:
                    cell_text = f"• {cell_text}"

                ax.text(x + 0.1, y_pos - 0.15,
                       cell_text, ha='left', va='top',
                       fontsize=9.5, color=COLORS['cell_text'],
                       fontweight='500',
                       wrap=True)

    # Draw range labels (column B) with arrows
    arrow_y_positions = []
    range_colors = {
        '$1M': COLORS['range_1m_bg'],
        '$500k': COLORS['range_500k_bg'],
        '$250k': COLORS['range_250k_bg']
    }

    for range_label, (start_row, end_row) in range_rows.items():
        # Calculate center position for the range
        start_y = 15.5 - (start_row * row_height)
        end_y = 15.5 - ((end_row + 1) * row_height)
        center_y = (start_y + end_y) / 2
        arrow_y_positions.append(center_y)

        # Draw range label with color-coded background
        ax.text(col_positions['B'], center_y,
               range_label, ha='center', va='center',
               fontsize=17, fontweight='bold', color=COLORS['range_text'],
               bbox=dict(boxstyle='round,pad=0.6',
                        facecolor=range_colors[range_label],
                        edgecolor='white', linewidth=3))

    # Draw arrows between the ranges
    from matplotlib.patches import FancyArrowPatch
    for i in range(len(arrow_y_positions) - 1):
        arrow = FancyArrowPatch(
            (col_positions['B'], arrow_y_positions[i] - 0.8),
            (col_positions['B'], arrow_y_positions[i+1] + 0.8),
            arrowstyle='->,head_width=0.5,head_length=0.5',
            color=COLORS['arrow_color'],
            linewidth=3,
            mutation_scale=25
        )
        ax.add_patch(arrow)

    # Legend (rows 15-16)
    legend_y = 15.5 - (15 * row_height)

    if 'B15' in cells:
        ax.text(col_positions['B'], legend_y - 0.3,
               cells['B15'], ha='left', va='center',
               fontsize=10, fontweight='bold', color=COLORS['cell_text'])

    if 'C15' in cells:
        ax.text(col_positions['C'], legend_y - 0.3,
               cells['C15'], ha='left', va='center',
               fontsize=9, color=COLORS['cell_text'],
               bbox=dict(boxstyle='round,pad=0.3',
                        facecolor='#E8F4F8',
                        edgecolor=COLORS['cell_border'], linewidth=1))

    if 'C16' in cells:
        ax.text(col_positions['C'], legend_y - 1.1,
               cells['C16'], ha='left', va='center',
               fontsize=9, color=COLORS['cell_text'],
               bbox=dict(boxstyle='round,pad=0.3',
                        facecolor='#FFF5E6',
                        edgecolor=COLORS['cell_border'], linewidth=1))

    plt.tight_layout()

    # Save
    os.makedirs('.tmp/charts', exist_ok=True)
    output_path = '.tmp/charts/savings_exact_positioning.png'
    plt.savefig(output_path, dpi=300, bbox_inches='tight',
               facecolor='white', edgecolor='none')
    print(f"\n✓ Chart saved with exact positioning: {output_path}")

    plt.close()
    return output_path

if __name__ == "__main__":
    create_exact_chart()
