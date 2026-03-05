"""
Analyze Excel file structure and extract chart data
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference
import os

def analyze_excel(excel_path):
    """
    Thoroughly analyze Excel file to understand structure and extract charts

    Args:
        excel_path: Path to Excel file
    """
    print(f"Analyzing: {excel_path}\n")

    # Read with openpyxl to access charts and formatting
    wb = load_workbook(excel_path)

    for sheet_name in wb.sheetnames:
        print(f"Sheet: {sheet_name}")
        ws = wb[sheet_name]

        # Print all non-empty cells
        print("\nAll cell values:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10), 1):
            row_values = []
            for col_idx, cell in enumerate(row, 1):
                if cell.value is not None and str(cell.value).strip():
                    row_values.append(f"  {chr(64+col_idx)}{row_idx}: {cell.value}")
            if row_values:
                print(f"Row {row_idx}:")
                for val in row_values:
                    print(val)

        # Check for charts
        print(f"\n\nCharts in sheet '{sheet_name}':")
        if hasattr(ws, '_charts'):
            print(f"  Found {len(ws._charts)} chart(s)")

            for idx, chart in enumerate(ws._charts, 1):
                print(f"\n  Chart #{idx}:")
                print(f"    Type: {type(chart).__name__}")
                print(f"    Title: {chart.title}")
                print(f"    Position: {chart.anchor}")

                # Try to extract data from chart
                if hasattr(chart, 'series'):
                    print(f"    Series count: {len(chart.series)}")
                    for s_idx, series in enumerate(chart.series, 1):
                        print(f"\n    Series #{s_idx}:")
                        if hasattr(series, 'title'):
                            print(f"      Title: {series.title}")
                        if hasattr(series, 'val'):
                            print(f"      Values: {series.val}")
                        if hasattr(series, 'cat'):
                            print(f"      Categories: {series.cat}")

    # Also read as DataFrame to see structure
    print("\n" + "=" * 80)
    print("\nDataFrame analysis:")
    try:
        df = pd.read_excel(excel_path, sheet_name=0)
        print("\nFull DataFrame (first 30 rows):")
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 30)
        print(df.head(30))

        print(f"\n\nNon-NaN cell locations:")
        for col in df.columns:
            non_nan = df[col].notna().sum()
            if non_nan > 0:
                print(f"  {col}: {non_nan} non-null values")
                non_null_vals = df[col].dropna().head(10)
                if len(non_null_vals) > 0:
                    print(f"    Sample values: {non_null_vals.tolist()}")

    except Exception as e:
        print(f"Error reading DataFrame: {e}")

if __name__ == "__main__":
    excel_path = '.tmp/email_attachments/Savings overview01.xlsx'
    if os.path.exists(excel_path):
        analyze_excel(excel_path)
    else:
        print(f"File not found: {excel_path}")
